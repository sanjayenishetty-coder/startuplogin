#!/usr/bin/env python3
"""Build data/startups.js from one or more raw startup lists (CSV or XLSX).

Usage:
    python3 scripts/transform_data.py <file1.csv> [file2.xlsx ...]

Columns are matched by header name (case/space tolerant), so sheets with
slightly different layouts all work. Recognised headers:
    name:        "Start up Name", "Startup/ Company name", "Startup / Company Name"
    founded:     "Founded Year", "Incorporation Year"
    industry:    "Industry", "Industry / Sector"
    stage:       "Stage"
    tagline:     "One Liner"
    description: "Description"
    website:     "Weblink", "Website"
    location:    "Location"
    founders:    "Founder", "Founder(s)"
    investors:   "Key Investors"
    funding:     "Total Funding", "Total Funding Received"

Privacy: contact details, LinkedIn URLs, revenue and internal rating/comment
columns are intentionally NOT exported to the public site.
"""
import csv
import hashlib
import json
import re
import sys

CITY_MAP = {
    "bengaluru": ("Bengaluru", "Karnataka", 12.9716, 77.5946),
    "bangalore": ("Bengaluru", "Karnataka", 12.9716, 77.5946),
    "hyderabad": ("Hyderabad", "Telangana", 17.4260, 78.4520),
    "secunderabad": ("Hyderabad", "Telangana", 17.4399, 78.4983),
    "mumbai": ("Mumbai", "Maharashtra", 19.0760, 72.8777),
    "thane": ("Mumbai", "Maharashtra", 19.2183, 72.9781),
    "navi mumbai": ("Mumbai", "Maharashtra", 19.0330, 73.0297),
    "pune": ("Pune", "Maharashtra", 18.5204, 73.8567),
    "delhi": ("Delhi", "Delhi NCR", 28.6139, 77.2090),
    "new delhi": ("Delhi", "Delhi NCR", 28.6139, 77.2090),
    "gurgaon": ("Gurugram", "Delhi NCR", 28.4595, 77.0266),
    "gurugram": ("Gurugram", "Delhi NCR", 28.4595, 77.0266),
    "haryana": ("Gurugram", "Delhi NCR", 28.4595, 77.0266),
    "noida": ("Noida", "Delhi NCR", 28.5355, 77.3910),
    "faridabad": ("Faridabad", "Delhi NCR", 28.4089, 77.3178),
    "ghaziabad": ("Ghaziabad", "Delhi NCR", 28.6692, 77.4538),
    "chennai": ("Chennai", "Tamil Nadu", 13.0827, 80.2707),
    "coimbatore": ("Coimbatore", "Tamil Nadu", 11.0168, 76.9558),
    "jaipur": ("Jaipur", "Rajasthan", 26.9124, 75.7873),
    "kota": ("Kota", "Rajasthan", 25.2138, 75.8648),
    "ahmedabad": ("Ahmedabad", "Gujarat", 23.0225, 72.5714),
    "surat": ("Surat", "Gujarat", 21.1702, 72.8311),
    "indore": ("Indore", "Madhya Pradesh", 22.7196, 75.8577),
    "bhopal": ("Bhopal", "Madhya Pradesh", 23.2599, 77.4126),
    "kolkata": ("Kolkata", "West Bengal", 22.5726, 88.3639),
    "kolkatta": ("Kolkata", "West Bengal", 22.5726, 88.3639),
    "howrah": ("Kolkata", "West Bengal", 22.5958, 88.2636),
    "lucknow": ("Lucknow", "Uttar Pradesh", 26.8467, 80.9462),
    "kanpur": ("Kanpur", "Uttar Pradesh", 26.4499, 80.3319),
    "bhubhaneshwar": ("Bhubaneswar", "Odisha", 20.2961, 85.8245),
    "bhubaneswar": ("Bhubaneswar", "Odisha", 20.2961, 85.8245),
    "roorkee": ("Roorkee", "Uttarakhand", 29.8543, 77.8880),
    "dehradun": ("Dehradun", "Uttarakhand", 30.3165, 78.0322),
    "ahmednagar": ("Ahmednagar", "Maharashtra", 19.0948, 74.7480),
    "nagpur": ("Nagpur", "Maharashtra", 21.1458, 79.0882),
    "nashik": ("Nashik", "Maharashtra", 19.9975, 73.7898),
    "chandigarh": ("Chandigarh", "Chandigarh", 30.7333, 76.7794),
    "mohali": ("Mohali", "Punjab", 30.7046, 76.7179),
    "kochi": ("Kochi", "Kerala", 9.9312, 76.2673),
    "thiruvananthapuram": ("Thiruvananthapuram", "Kerala", 8.5241, 76.9366),
    "trivandrum": ("Thiruvananthapuram", "Kerala", 8.5241, 76.9366),
    "visakhapatnam": ("Visakhapatnam", "Andhra Pradesh", 17.6868, 83.2185),
    "vijayawada": ("Vijayawada", "Andhra Pradesh", 16.5062, 80.6480),
    "guwahati": ("Guwahati", "Assam", 26.1445, 91.7362),
    "patna": ("Patna", "Bihar", 25.5941, 85.1376),
    "ranchi": ("Ranchi", "Jharkhand", 23.3441, 85.3096),
    "goa": ("Goa", "Goa", 15.4909, 73.8278),
    "panaji": ("Goa", "Goa", 15.4909, 73.8278),
    "boston": ("Boston", "Outside India", 42.3601, -71.0589),
}

STAGE_MAP = {
    "pre seed": "Pre-seed", "preseed": "Pre-seed", "pre-seed": "Pre-seed",
    "seed": "Seed", "early stage": "Seed",
    "series a": "Series A", "series b": "Series B", "series c": "Series C+",
    "seriesc": "Series C+", "series c+": "Series C+", "seriesd": "Series C+",
    "series d": "Series C+", "series e": "Series C+",
    "unfunded": "Bootstrapped", "bootstrapped": "Bootstrapped",
    "public": "Public", "acquired": "Acquired",
}

# keyword (lowercase substring of raw industry / one-liner) -> sector
SECTOR_RULES = [
    ("agri", "Agritech"), ("farm", "Agritech"), ("croptech", "Agritech"),
    ("crop tech", "Agritech"), ("food waste", "Agritech"),
    ("fintech", "Fintech"), ("financial", "Fintech"), ("lending", "Fintech"),
    ("insur", "Fintech"), ("payment", "Fintech"), ("banking", "Fintech"),
    ("edtech", "Edtech"), ("education", "Edtech"), ("e-learning", "Edtech"),
    ("learning", "Edtech"),
    ("health", "Healthtech"), ("medical", "Healthtech"), ("pharma", "Healthtech"),
    ("hospital", "Healthtech"), ("wellness", "Healthtech"), ("biotech", "Healthtech"),
    ("disease", "Healthtech"), ("mental", "Healthtech"),
    ("game", "Gaming"), ("gaming", "Gaming"), ("esports", "Gaming"),
    ("sports", "Sports"), ("spectator", "Sports"), ("fitness", "Sports"),
    ("renewable", "Cleantech"), ("solar", "Cleantech"), ("climate", "Cleantech"),
    ("energy", "Cleantech"), ("environment", "Cleantech"), ("recycl", "Cleantech"),
    ("waste", "Cleantech"), ("electric vehicle", "Cleantech"),
    ("scm", "Logistics"), ("logistic", "Logistics"), ("supply chain", "Logistics"),
    ("warehouse", "Logistics"), ("shipping", "Logistics"), ("transport", "Logistics"),
    ("fleet", "Logistics"),
    ("hrtech", "HRtech"), ("hr tech", "HRtech"), ("recruit", "HRtech"),
    ("staffing", "HRtech"), ("applicant", "HRtech"), ("talent", "HRtech"),
    ("employee", "HRtech"),
    ("vr/ar", "Deeptech"), ("robotic", "Deeptech"), ("drone", "Deeptech"),
    ("space", "Spacetech"), ("semiconductor", "Deeptech"), ("iot", "Deeptech"),
    ("deep tech", "Deeptech"), ("deeptech", "Deeptech"), ("3d print", "Deeptech"),
    ("cyber", "Cybersecurity"), ("security", "Cybersecurity"),
    ("artificial intelligence", "AI"), (" ai", "AI"), ("ai ", "AI"),
    ("machine learning", "AI"), ("analytics", "AI"), ("big data", "AI"),
    ("data science", "AI"), ("computer vision", "AI"), ("conversational", "AI"),
    ("data visualisation", "AI"), ("data visualization", "AI"),
    ("e-commerce", "E-commerce"), ("ecommerce", "E-commerce"), ("marketplace", "E-commerce"),
    ("retail", "E-commerce"), ("grocery", "E-commerce"), ("d2c", "D2C"),
    ("consumer goods", "D2C"), ("fashion", "D2C"), ("beauty", "D2C"), ("apparel", "D2C"),
    ("food", "Foodtech"), ("beverage", "Foodtech"), ("restaurant", "Foodtech"),
    ("advertis", "Media & Adtech"), ("marketing", "Media & Adtech"),
    ("media", "Media & Adtech"), ("entertainment", "Media & Adtech"),
    ("publishing", "Media & Adtech"), ("content", "Media & Adtech"),
    ("real estate", "Proptech"), ("proptech", "Proptech"), ("construction", "Proptech"),
    ("travel", "Traveltech"), ("hospitality", "Traveltech"), ("tourism", "Traveltech"),
    ("legal", "Legaltech"),
    ("saas", "SaaS"), ("software development", "SaaS"), ("software", "SaaS"),
    ("productivity", "SaaS"), ("crm", "SaaS"), ("cloud", "SaaS"),
    ("it services", "IT Services"), ("it consulting", "IT Services"),
    ("information technology", "IT Services"), ("information services", "IT Services"),
    ("manufactur", "Manufacturing"), ("machinery", "Manufacturing"),
    ("electronics", "Manufacturing"), ("automotive", "Manufacturing"),
    ("internet", "Consumer"), ("technology", "SaaS"), ("telecom", "SaaS"),
]

HEADER_ALIASES = {
    "name": ["start up name", "startup/ company name", "startup / company name",
             "startup name", "company name", "name"],
    "founded": ["founded year", "incorporation year", "founded"],
    "industry": ["industry / sector", "industry"],
    "stage": ["stage"],
    "tagline": ["one liner", "one-line description", "oneliner"],
    "description": ["description"],
    "website": ["weblink", "website"],
    "location": ["location", "city"],
    "founders": ["founder(s)", "founder", "founders"],
    "investors": ["key investors", "investors"],
    "funding": ["total funding received", "total funding", "funding"],
}


def map_headers(header_row):
    idx = {}
    normed = [re.sub(r"\s+", " ", str(h or "").strip().lower()) for h in header_row]
    for field, aliases in HEADER_ALIASES.items():
        for a in aliases:
            if a in normed:
                idx[field] = normed.index(a)
                break
    return idx


def norm_stage(raw):
    s = re.sub(r"\s+", " ", (raw or "").strip().lower())
    if s in STAGE_MAP:
        return STAGE_MAP[s]
    if "series" in s:
        for k, v in STAGE_MAP.items():
            if k in s:
                return v
    if s.startswith("$") or "raised" in s:
        return "Seed"
    return ""


def norm_sector(industry, one_liner, description):
    text = " " + " ".join(filter(None, [industry, one_liner])).lower() + " "
    for kw, sector in SECTOR_RULES:
        if kw in text:
            return sector
    text = " " + (description or "").lower() + " "
    for kw, sector in SECTOR_RULES:
        if kw in text:
            return sector
    return "Others"


def jitter(name, lat, lng, spread=0.055):
    """Deterministic pseudo-random offset so city pins spread out stably."""
    h = hashlib.md5(name.encode("utf-8")).digest()
    dx = (int.from_bytes(h[0:4], "big") / 2**32 - 0.5) * 2 * spread
    dy = (int.from_bytes(h[4:8], "big") / 2**32 - 0.5) * 2 * spread
    return round(lat + dy, 5), round(lng + dx, 5)


def slugify(name):
    s = re.sub(r"[^a-z0-9]+", "-", name.lower()).strip("-")
    return s or "startup"


def clean_url(u):
    u = re.sub(r"\s+", "", str(u or ""))
    if not u or u.upper() == "NA":
        return ""
    if not u.startswith("http"):
        u = "https://" + u
    return u


def clean_founders(raw):
    """Keep only person names; strip emails/phones and role suffix noise."""
    if not raw or str(raw).strip().upper() == "NA":
        return ""
    parts = re.split(r"[,;\n]", str(raw))
    names = []
    for p in parts:
        p = p.strip()
        if not p or "@" in p or "linkedin" in p.lower() or re.search(r"\d{5,}", p):
            continue
        if re.fullmatch(r"(?i)(founder|co[- ]?founder|ceo|cto|coo|cpo|md|director|chairman|managing director|president)[.\s]*", p):
            continue
        p = re.sub(r"(?i)\s*\b(co[- ]?founder|founder|ceo|cto|coo|cpo|md|director)\b[.\s]*$", "", p).strip()
        if p and len(p) > 2:
            names.append(p)
    seen, out = set(), []
    for n in names:
        if n.lower() not in seen:
            seen.add(n.lower())
            out.append(n)
    return ", ".join(out[:6])


def clean_text(t, limit=None):
    t = re.sub(r"\s+", " ", str(t or "")).strip()
    if t.upper() == "NA":
        return ""
    if limit and len(t) > limit:
        t = t[: limit - 1].rsplit(" ", 1)[0] + "…"
    return t


def tidy_tagline(t):
    """One crisp line: prefer the first sentence when the raw tagline rambles."""
    t = clean_text(t)
    if len(t) > 110:
        first = re.split(r"(?<=[.!?])\s+", t)[0].strip()
        if 30 <= len(first) <= 130:
            t = first
        else:
            t = clean_text(t, 120)
    return t.rstrip(".")


def iter_rows(path):
    if path.lower().endswith((".xlsx", ".xlsm")):
        import openpyxl
        ws = openpyxl.load_workbook(path, read_only=True).active
        for row in ws.iter_rows(values_only=True):
            yield ["" if c is None else str(c) for c in row]
    else:
        for row in csv.reader(open(path, encoding="utf-8")):
            yield row


def parse_file(path):
    rows = iter_rows(path)
    header = next(rows)
    idx = map_headers(header)
    if "name" not in idx:
        raise SystemExit(f"{path}: could not find a name column in {header}")
    get = lambda r, f: r[idx[f]] if f in idx and idx[f] < len(r) else ""
    out = []
    for r in rows:
        name = clean_text(get(r, "name"))
        if not name:
            continue
        founded = re.sub(r"\D", "", clean_text(get(r, "founded")))[:4]
        loc_key = re.sub(r"\s+", " ", clean_text(get(r, "location")).lower())
        city, state, lat, lng = CITY_MAP.get(loc_key, ("", "", None, None))
        if not city and loc_key and loc_key != "na":
            city, state = clean_text(get(r, "location")).title(), "India"
        entry = {
            "name": name,
            "type": "startup",
            "tagline": tidy_tagline(get(r, "tagline")),
            "description": clean_text(get(r, "description"), 700),
            "website": clean_url(get(r, "website")),
            "city": city,
            "state": state,
            "sector": norm_sector(get(r, "industry"), get(r, "tagline"), get(r, "description")),
            "industry": clean_text(get(r, "industry"), 90).rstrip("."),
            "stage": norm_stage(get(r, "stage")),
            "founded": founded if founded and 1990 <= int(founded or 0) <= 2026 else "",
            "founders": clean_founders(get(r, "founders")),
            "investors": clean_text(get(r, "investors"), 140),
            "funding": clean_text(get(r, "funding"), 30),
        }
        if lat is not None:
            entry["lat"], entry["lng"] = jitter(name, lat, lng)
        out.append(entry)
    return out


def main(paths):
    merged, by_key = [], {}
    for path in paths:
        entries = parse_file(path)
        added = 0
        for e in entries:
            key = re.sub(r"[^a-z0-9]", "", e["name"].lower())
            if key in by_key:
                # merge: fill blanks in the existing record
                old = by_key[key]
                for f, v in e.items():
                    if v and not old.get(f):
                        old[f] = v
                continue
            by_key[key] = e
            merged.append(e)
            added += 1
        print(f"{path}: {len(entries)} rows, {added} new")

    slugs = set()
    for e in merged:
        slug = slugify(e["name"])
        while slug in slugs:
            slug += "-2"
        slugs.add(slug)
        e["slug"] = slug

    js = ("// Auto-generated by scripts/transform_data.py — do not edit by hand.\n"
          "// Regenerate: python3 scripts/transform_data.py <sheet1.csv> [sheet2.xlsx ...]\n"
          "window.STARTUP_DATA = " + json.dumps(merged, ensure_ascii=False, indent=1) + ";\n")
    with open("data/startups.js", "w", encoding="utf-8") as f:
        f.write(js)
    cities = {}
    for e in merged:
        cities[e["city"]] = cities.get(e["city"], 0) + 1
    print(f"wrote data/startups.js: {len(merged)} startups, {len(cities)} cities")
    print(sorted(cities.items(), key=lambda x: -x[1])[:15])


if __name__ == "__main__":
    main(sys.argv[1:])
